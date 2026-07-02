#!/usr/bin/env python3
"""Parse the EI Statistical Review workbook into a clean JSON dataset for the dashboard."""
import openpyxl, json, re, math

SRC = "/sessions/zealous-beautiful-hypatia/mnt/uploads/EI-Stats-Review-ALL-data (2).xlsx"
OUT = "/sessions/zealous-beautiful-hypatia/mnt/outputs/ei_data.json"

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

# ---------- helpers ----------
def is_year(v):
    if isinstance(v, int) and 1850 <= v <= 2035:
        return True
    if isinstance(v, float) and v.is_integer() and 1850 <= v <= 2035:
        return True
    return False

def clean_num(v):
    """Return float (rounded) or None for non-numeric / na markers."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        # round to 6 sig figs to keep file small but lossless-enough
        if v == 0:
            return 0
        try:
            r = round(v, 6 - int(math.floor(math.log10(abs(v)))) - 1)
        except Exception:
            r = v
        if isinstance(r, float) and r.is_integer():
            return int(r)
        return r
    s = str(v).strip()
    if s in ("-", "–", "—", "", "n/a", "N/A", "na", "NA", "^", "*"):
        return None
    s2 = s.replace(",", "")
    try:
        f = float(s2)
        return int(f) if f.is_integer() else round(f, 4)
    except Exception:
        return None

def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()

# Merge clearly-identical entities that the workbook spells differently across sheets.
ALIASES = {
    "United States": "US",
    "Russia": "Russian Federation",
    "Turkey": "Türkiye",
    "Viet Nam": "Vietnam",
    "Czechia": "Czech Republic",
    "UAE": "United Arab Emirates",
    "DR Congo": "Democratic Republic of Congo",
    "European Union #": "European Union",
    "Total Asia-Pacific": "Total Asia Pacific",
}

def norm_name(s):
    """Normalize an entity / series name: collapse whitespace, fix the one source typo,
    strip trailing footnote markers (e.g. 'India2' -> 'India', 'European Union #' -> 'European Union'),
    and merge known duplicate spellings to a single canonical name."""
    s = re.sub(r"\s+", " ", str(s)).strip()
    s = s.replace("Russian Federationeration", "Russian Federation")
    # strip trailing superscript / footnote markers (incl. '#')
    s = re.sub(r"[¹²³⁰-⁹\*\^#]+$", "", s).strip()
    # strip a trailing footnote digit run if preceded by a letter (country names don't end in digits)
    m = re.match(r"^(.*[A-Za-z\)])\s?(\d{1,2})$", s)
    if m:
        s = m.group(1).strip()
    return ALIASES.get(s, s)

AGG_PATTERNS = [
    r"^total\b", r"\btotal$", r"^of which", r"\boecd\b", r"non-oecd",
    r"european union", r"^eu\b", r"\bworld\b", r"^other\b", r"opec",
    r"cis$", r"^cis\b", r"rest of", r"middle east$", r"africa$",
    r"asia pacific$", r"europe$", r"north america$",
    r"s\. & cent\. america", r"central america$",
]
def is_aggregate(name):
    n = name.lower().strip()
    for p in AGG_PATTERNS:
        if re.search(p, n):
            return True
    return False

# ---------- category mapping ----------
def categorize(name):
    n = name.lower()
    if name in ("Contents", "Definitions", "Approximate conversion factors"):
        return "Reference"
    if "carbon price" in n or "co2" in n or "co2e" in n or "ccus" in n or "flaring" in n or "methane" in n:
        return "Carbon & Emissions"
    if (name.startswith("Oil") or "oil" in n.split() or "crude" in n or "spot crude" in n
            or n.startswith("ngls") or n.startswith("liquids") or "refin" in n):
        return "Oil"
    if name.startswith("Gas") or n.startswith("gas "):
        return "Natural Gas"
    if name.startswith("Coal"):
        return "Coal"
    if "nuclear" in n:
        return "Nuclear"
    if "hydro" in n:
        return "Hydro"
    if any(k in n for k in ["renewable", "solar", "wind", "geo biomass", "biofuel", "saf"]):
        return "Renewables"
    if "elec" in n:
        return "Electricity"
    if "bess" in n or "data centre" in n or "data center" in n:
        return "Storage & Demand"
    if any(k in n for k in ["cobalt", "lithium", "graphite", "rare earth", "copper", "manganese", "nickel", "zinc", "mineral commodity"]):
        return "Critical Minerals"
    if "total energy" in n or "tes" in n:
        return "Overview"
    return "Other"

CATEGORY_ORDER = ["Derived & ratios", "Overview", "Carbon & Emissions", "Oil", "Natural Gas", "Coal",
                  "Nuclear", "Hydro", "Renewables", "Electricity", "Storage & Demand",
                  "Critical Minerals", "Other", "Reference"]

# ---------- per-sheet parse ----------
def load_grid(ws):
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(list(r))
    # trim fully-empty trailing rows/cols
    while rows and all(c is None for c in rows[-1]):
        rows.pop()
    if not rows:
        return []
    maxc = max((max((i for i, c in enumerate(r) if c is not None), default=-1) for r in rows), default=-1) + 1
    rows = [r[:maxc] for r in rows]
    return rows

def find_year_header_row(grid):
    """Return (row_idx, [(col_idx, year)...]) for the row with the most year cells, or None."""
    best = None
    for ri, row in enumerate(grid[:8]):
        yrs = [(ci, int(c)) for ci, c in enumerate(row) if is_year(c)]
        # require at least 2 years and that they are increasing-ish
        if len(yrs) >= 2:
            if best is None or len(yrs) > len(best[1]):
                best = (ri, yrs)
    return best

def parse_sheet(name, ws):
    grid = load_grid(ws)
    cat = categorize(name)
    title = ""
    for row in grid[:3]:
        for c in row:
            if c is not None and str(c).strip():
                title = str(c).strip()
                break
        if title:
            break

    out = {"name": name, "category": cat, "title": title, "type": "raw"}

    # Reference / contents -> raw only
    if cat == "Reference":
        out["grid"] = [[cell_str(c) for c in row] for row in grid]
        return out

    yh = find_year_header_row(grid)

    # ---- Year-indexed (prices): years run DOWN the first column ----
    col0_years = [r for r in range(len(grid)) if grid[r] and is_year(grid[r][0])]
    header_for_series = None
    if len(col0_years) >= 4 and (yh is None or len(col0_years) > len(yh[1])):
        # series headers: the last non-empty text row above first year row
        first_yr_row = col0_years[0]
        hdr_row_idx = None
        for r in range(first_yr_row - 1, -1, -1):
            texts = [ci for ci, c in enumerate(grid[r]) if c is not None and str(c).strip() and not is_year(c)]
            if len(texts) >= 2:
                hdr_row_idx = r
                break
        series = []
        ncols = max(len(grid[r]) for r in col0_years)
        for ci in range(1, ncols):
            label = ""
            if hdr_row_idx is not None and ci < len(grid[hdr_row_idx]):
                label = cell_str(grid[hdr_row_idx][ci])
            # combine multi-row header if needed
            if not label:
                for r in range(max(0, first_yr_row - 4), first_yr_row):
                    if ci < len(grid[r]) and cell_str(grid[r][ci]):
                        label = cell_str(grid[r][ci]); break
            series.append(label or f"Series {ci}")
        years = [int(grid[r][0]) for r in col0_years]
        seriesdata = []
        for ci in range(1, ncols):
            vals = []
            for r in col0_years:
                v = grid[r][ci] if ci < len(grid[r]) else None
                vals.append(clean_num(v))
            if any(v is not None for v in vals):
                seriesdata.append({"name": norm_name(series[ci-1]), "values": vals})
        unit = ""
        for r in range(first_yr_row):
            for c in grid[r]:
                if c and ("per " in str(c).lower() or "dollar" in str(c).lower() or "$" in str(c)):
                    unit = str(c).strip(); break
            if unit: break
        out.update({"type": "yearindexed", "years": years, "series": seriesdata, "unit": unit})
        out["grid"] = [[cell_str(c) for c in row] for row in grid]
        return out

    # ---- Standard timeseries: countries x years ----
    if yh is not None and len(yh[1]) >= 3:
        hr, yrs = yh
        # Keep ONLY the contiguous, strictly-increasing run of year columns starting
        # at the first year cell. Trailing cols whose header is also a year (e.g. a
        # "Growth rate ... 2025" or "Share 2025" column) break the run and become extras.
        run = [yrs[0]]
        for ci, y in yrs[1:]:
            pci, py = run[-1]
            if ci == pci + 1 and y == py + 1:
                run.append((ci, y))
            else:
                break
        year_cols = run
        years = [y for ci, y in year_cols]
        # Guard: a "time series" with fewer than 3 real years isn't a usable chart
        # (e.g. "Wind and Solar Technology Types" is a 2024/2025 cross-section, not a
        # trend). Fall through to the raw/custom-layout renderer instead.
        if len(years) < 3:
            out["grid"] = [[cell_str(c) for c in row] for row in grid]
            return out
        ycols = [ci for ci, y in year_cols]
        first_ycol = min(ycols)
        last_ycol = max(ycols)
        # unit = text in header row before years, or row above
        unit = ""
        for ci in range(first_ycol):
            if hr < len(grid) and ci < len(grid[hr]) and cell_str(grid[hr][ci]):
                unit = cell_str(grid[hr][ci]); break
        if not unit:
            for r in range(hr):
                if grid[r] and cell_str(grid[r][0]):
                    t = cell_str(grid[r][0])
                    if any(u in t.lower() for u in ["joule","tonne","barrel","cubic","watt","capita","megawatt","percent","%","kbo","mt","bcm","bcf","twh","ej","pj","gj"]):
                        unit = t; break
        # trailing extra columns (growth/share) - headers from hr row after last year col.
        # The descriptor (e.g. "Growth rate per annum") often sits in the row above and is
        # a merged cell, so carry it forward left-to-right across the extra region.
        ncol_total = max((len(grid[r]) for r in range(hr, len(grid))), default=0)
        extra_cols = []
        carry = ""
        for ci in range(last_ycol + 1, ncol_total):
            lbl = cell_str(grid[hr][ci]) if ci < len(grid[hr]) else ""
            above = ""
            for r in range(max(0, hr-2), hr):
                if ci < len(grid[r]) and cell_str(grid[r][ci]):
                    above = cell_str(grid[r][ci])
            if above:
                carry = above
            full = (carry + " " + lbl).strip() if carry else lbl
            full = re.sub(r"\s+", " ", full).strip()
            if full:
                extra_cols.append({"col": ci, "name": full})

        entities = []
        notes = []
        data_start = hr + 1
        for r in range(data_start, len(grid)):
            row = grid[r]
            label = cell_str(row[0]) if row else ""
            # detect footnotes (label starts with note markers or appears after data with no numbers)
            vals = []
            for ci in ycols:
                v = row[ci] if ci < len(row) else None
                vals.append(clean_num(v))
            has_num = any(v is not None for v in vals)
            if not label and not has_num:
                if entities:
                    entities[-1]["gapAfter"] = True
                continue
            if label and not has_num:
                # could be a footnote / note row, or a sub-header. Heuristic:
                low = label.lower()
                if (label[0] in "*^123456789♦" or low.startswith(("note","source","*","w/o","less than","n/a"))
                        or len(label) > 60):
                    notes.append(re.sub(r"\s+"," ", label).strip())
                    continue
                # else treat as a blank-data entity (rare) - skip
                continue
            extras = {}
            for ec in extra_cols:
                ev = row[ec["col"]] if ec["col"] < len(row) else None
                cv = clean_num(ev)
                if cv is not None:
                    extras[ec["name"]] = cv
            ent = {"name": norm_name(label), "values": vals}
            if is_aggregate(ent["name"]):
                ent["agg"] = True
            if extras:
                ent["extra"] = extras
            entities.append(ent)

        out.update({"type": "timeseries", "years": years, "unit": re.sub(r"\s+"," ", unit).strip(),
                    "entities": entities,
                    "extraCols": [e["name"] for e in extra_cols],
                    "notes": notes})
        return out

    # ---- fallback: raw grid (trade matrices, by-fuel, etc.) ----
    out["grid"] = [[cell_str(c) for c in row] for row in grid]
    return out

# ---------- run ----------
sheets = {}
order = []
for name in wb.sheetnames:
    ws = wb[name]
    try:
        parsed = parse_sheet(name, ws)
    except Exception as e:
        parsed = {"name": name, "category": categorize(name), "type": "raw",
                  "grid": [[cell_str(c) for c in row] for row in load_grid(ws)], "error": str(e)}
    sheets[name] = parsed
    order.append(name)

# ---------- DERIVED & RATIOS: synthetic cross-table analyses ----------
def build_derived(sheets):
    def grab(name):
        """sheet -> (dict ent->{year:val}, years, dict ent->agg, ordered entity list)."""
        s = sheets[name]
        data = {e["name"]: dict(zip(s["years"], e["values"])) for e in s["entities"]}
        agg = {e["name"]: bool(e.get("agg")) for e in s["entities"]}
        return data, s["years"], agg, [e["name"] for e in s["entities"]]

    out = {}
    def mk(name, title, unit, years, ref_order, ref_agg, compute, note=""):
        ents = []
        for en in ref_order:
            vals = []
            has = False
            for y in years:
                v = compute(en, y)
                if v is not None and not (isinstance(v, float) and (v != v)):
                    v = round(v, 3); has = True
                else:
                    v = None
                vals.append(v)
            if has:
                e = {"name": en, "values": vals}
                if ref_agg.get(en):
                    e["agg"] = True
                ents.append(e)
        out[name] = {"name": name, "category": "Derived & ratios", "title": title,
                     "unit": unit, "type": "timeseries", "years": list(years),
                     "entities": ents, "extraCols": [], "notes": [note] if note else []}

    EG, egY, egAgg, egOrder = grab("Electricity Generation - TWh")
    NUC = grab("Nuclear Generation - TWh")[0]; HYD = grab("Hydro Generation - TWh")[0]
    SOL = grab("Solar Generation - TWh")[0]; WND = grab("Wind Generation - TWh")[0]
    GEO = grab("Geo Biomass Other - TWh")[0]
    TES, tesY, tesAgg, tesOrder = grab("Total Energy Supply (TES) -EJ")
    PC = grab("TES per Capita")[0]
    OILc = grab("Oil Consumption - EJ")[0]; GASc = grab("Gas Consumption - EJ")[0]; COALc = grab("Coal Consumption - EJ")[0]
    CO2 = grab("CO2 from Energy")[0]
    OILp = grab("Oil Production - barrels")[0]; OILcb = grab("Oil Consumption - barrels")[0]
    GASp = grab("Gas Production - Bcm")[0]; GAScb = grab("Gas Consumption - Bcm")[0]
    COALg = grab("Coal inputs - Elec generation ")[0]; GASg = grab("Gas inputs - Elec generation")[0]

    def g(d, en, y):
        v = d.get(en, {}).get(y); return v if isinstance(v, (int, float)) else None

    # 1) Clean (non-fossil) electricity share
    def clean_share(en, y):
        tot = g(EG, en, y)
        if not tot: return None
        cl = sum(x for x in (g(NUC,en,y),g(HYD,en,y),g(SOL,en,y),g(WND,en,y),g(GEO,en,y)) if x is not None)
        return cl / tot * 100
    mk("Clean electricity share", "Non-fossil generation as % of total electricity", "%",
       egY, egOrder, egAgg, clean_share,
       "Clean = nuclear + hydro + solar + wind + geo/biomass/other, divided by total generation.")

    # 2) Wind + solar share of electricity
    def ws_share(en, y):
        tot = g(EG, en, y)
        if not tot: return None
        ws = sum(x for x in (g(SOL,en,y),g(WND,en,y)) if x is not None)
        return ws / tot * 100
    mk("Wind & solar share of electricity", "Wind + solar as % of total electricity", "%",
       egY, egOrder, egAgg, ws_share)

    # 2b) Nuclear / coal / gas share of electricity (fuel mix companions)
    def fuel_share(src):
        def f(en, y):
            tot = g(EG, en, y)
            if not tot: return None
            return (g(src, en, y) or 0) / tot * 100
        return f
    mk("Nuclear share of electricity", "Nuclear as % of total electricity", "%",
       egY, egOrder, egAgg, fuel_share(NUC))
    mk("Coal share of electricity", "Coal-fired as % of total electricity", "%",
       egY, egOrder, egAgg, fuel_share(COALg),
       "Coal-fired generation ÷ total generation.")
    mk("Gas share of electricity", "Gas-fired as % of total electricity", "%",
       egY, egOrder, egAgg, fuel_share(GASg),
       "Gas-fired generation ÷ total generation.")

    # 3) Electrification rate: electricity as % of primary energy
    def electrify(en, y):
        tot = g(EG, en, y); tes = g(TES, en, y)
        if not tot or not tes: return None
        return (tot * 0.0036) / tes * 100   # TWh -> EJ
    mk("Electrification rate", "Electricity generation as % of primary energy (input-equivalent)", "%",
       [y for y in egY if y in tesY], egOrder, egAgg, electrify,
       "Uses EI input-equivalent primary energy; a final-energy basis would read a few points higher.")

    # 4) Fossil fuel share of primary energy
    def fossil_share(en, y):
        tes = g(TES, en, y)
        if not tes: return None
        f = sum(x for x in (g(OILc,en,y),g(GASc,en,y),g(COALc,en,y)) if x is not None)
        return f / tes * 100
    mk("Fossil share of primary energy", "Oil + gas + coal as % of total energy supply", "%",
       tesY, tesOrder, tesAgg, fossil_share)

    # 5) Carbon intensity of energy: CO2 / TES  (Mt/EJ == kg CO2 / GJ)
    def carb_int(en, y):
        co2 = g(CO2, en, y); tes = g(TES, en, y)
        if not co2 or not tes: return None
        return co2 / tes
    mk("Carbon intensity of energy", "CO₂ from energy per unit of primary energy", "kg CO₂ / GJ",
       tesY, tesOrder, tesAgg, carb_int,
       "kg CO₂ per gigajoule (= Mt/EJ = t/TJ). Lower is cleaner energy per unit consumed.")

    # 6) CO2 emissions per capita (derived population = TES / TES-per-capita)
    def co2_pc(en, y):
        co2 = g(CO2, en, y); tes = g(TES, en, y); pc = g(PC, en, y)
        if not co2 or not tes or not pc: return None
        pop = tes * 1e9 / pc                # people
        return co2 * 1e6 / pop              # t CO2 per person
    mk("CO₂ emissions per capita", "Energy CO₂ per person (derived population)", "t CO₂ / person",
       tesY, tesOrder, tesAgg, co2_pc,
       "Population derived as primary energy ÷ energy-per-capita; territorial (production) emissions.")

    # 7) Oil self-sufficiency (production / consumption), barrels basis
    def oil_ss(en, y):
        c = g(OILcb, en, y)
        if not c: return None
        p = g(OILp, en, y) or 0   # absent from the producers table = ~0 production
        return p / c * 100
    mk("Oil self-sufficiency", "Oil production as % of consumption (>100% = net exporter)", "%",
       grab("Oil Consumption - barrels")[1], grab("Oil Consumption - barrels")[3], grab("Oil Consumption - barrels")[2], oil_ss,
       "Above 100% = produces more than it uses (net exporter); below = import-dependent.")

    # 8) Gas self-sufficiency (production / consumption), Bcm basis
    def gas_ss(en, y):
        c = g(GAScb, en, y)
        if not c: return None
        p = g(GASp, en, y) or 0
        return p / c * 100
    mk("Gas self-sufficiency", "Gas production as % of consumption (>100% = net exporter)", "%",
       grab("Gas Consumption - Bcm")[1], grab("Gas Consumption - Bcm")[3], grab("Gas Consumption - Bcm")[2], gas_ss,
       "Above 100% = net exporter; below = import-dependent.")

    # 9) Data centres as % of electricity generation (regions; special denominators for the composite rows)
    DC, dcY, dcAgg, dcOrder = grab("Data Centre Demand")
    def dc_share(en, y):
        num = g(DC, en, y)
        if num is None: return None
        if en == "Other North America":
            den = (g(EG,"Total North America",y) or 0) - (g(EG,"US",y) or 0)
        elif en == "Total Middle East and Africa":
            den = (g(EG,"Total Middle East",y) or 0) + (g(EG,"Total Africa",y) or 0)
        elif en == "Other Asia Pacific":
            den = (g(EG,"Total Asia Pacific",y) or 0) - (g(EG,"China",y) or 0)
        else:
            den = g(EG, en, y)
        return num / den * 100 if den else None
    mk("Data centres % of electricity", "Data-centre demand as % of regional electricity generation", "%",
       dcY, dcOrder, dcAgg, dc_share,
       "Denominator is electricity generation (proxy for consumption). Composite regions use summed/derived denominators.")

    return out

derived = build_derived(sheets)
for nm, obj in derived.items():
    sheets[nm] = obj
    order.append(nm)

# group categories in order
cats = {}
for name in order:
    c = sheets[name]["category"]
    cats.setdefault(c, []).append(name)
categories = [{"name": c, "sheets": cats[c]} for c in CATEGORY_ORDER if c in cats]

# ---------- curated entity list for the Country Profile picker ----------
def build_profile(sheets):
    import unicodedata
    names = set()
    for s in sheets.values():
        if s["type"] == "timeseries":
            for e in s["entities"]:
                names.add(e["name"])

    # Major regions / groups worth profiling: (canonical entity name, friendly label)
    REGION_MAP = [
        ("Total World", "World"),
        ("of which: OECD", "OECD"),
        ("Non-OECD", "Non-OECD"),
        ("OPEC", "OPEC"),
        ("Non-OPEC", "Non-OPEC"),
        ("European Union", "European Union"),
        ("Total North America", "North America"),
        ("Total S. & Cent. America", "S. & Cent. America"),
        ("Total Europe", "Europe"),
        ("Total CIS", "CIS"),
        ("Total Middle East", "Middle East"),
        ("Total Africa", "Africa"),
        ("Total Asia Pacific", "Asia Pacific"),
    ]
    regions = [{"value": v, "label": lab} for v, lab in REGION_MAP if v in names]

    # things that are NOT a geography (trade flows, oil products, tech/meter labels, footnotes)
    JUNK_TOKENS = re.compile(
        r"import|export|\btrade\b|pipeline|distillate|naphtha|gasoline|kerosene|"
        r"diesel|gasoil|ethane|\blpg\b|behind-the-meter|front-of-the-meter", re.I)
    JUNK_EXACT = {"Others", "Other", "Blue", "Green", "Fuel oil",
                  "Light distillates", "Middle distillates",
                  "Behind-the-meter", "Front-of-the-meter"}
    # bare regions / sub-regions that aren't countries (and aren't in the curated region group)
    REGION_DROP = {
        "Europe", "Africa", "Asia Pacific", "Middle East", "CIS", "World",
        "North Africa", "West Africa", "Central America", "Eastern Africa",
        "Middle Africa", "Western Africa", "Northern Africa", "Southern Africa",
        "Canada & Mexico", "Asia Pacific (ex Japan)", "Middle East (ex Saudi Arabia)",
        "S. & Cent. America", "OECD Asia", "Non-OECD", "OPEC", "Non-OPEC",
        "European Union", "of which: OECD",
    }

    def is_country(n):
        if ":" in n or n.startswith("Source"):
            return False
        if n in JUNK_EXACT or JUNK_TOKENS.search(n):
            return False
        if n.startswith(("Total ", "Other ", "Rest of")):
            return False
        if n in REGION_DROP:
            return False
        return True

    def sortkey(n):
        k = unicodedata.normalize("NFKD", n).encode("ascii", "ignore").decode().lower()
        if k.startswith("the "):
            k = k[4:]
        return k

    countries = sorted([n for n in names if is_country(n)], key=sortkey)
    return {"regions": regions, "countries": [{"value": n, "label": n} for n in countries]}

profile = build_profile(sheets)

data = {
    "meta": {
        "title": "Energy Institute Statistical Review of World Energy 2026",
        "source": "Energy Institute Statistical Review of World Energy (2026 edition)",
        "sheetCount": len(order),
    },
    "categories": categories,
    "order": order,
    "sheets": sheets,
    "profile": profile,
}

with open(OUT, "w") as f:
    json.dump(data, f, separators=(",", ":"), ensure_ascii=False)

import os
size = os.path.getsize(OUT)
print(f"Wrote {OUT}  ({size/1e6:.2f} MB)")

# summary of types
from collections import Counter
tc = Counter(s["type"] for s in sheets.values())
print("Types:", dict(tc))
print("Categories:", {c["name"]: len(c["sheets"]) for c in categories})
# sanity: list timeseries sheets with year span + entity count
print("\nTimeseries sheets (sample):")
for name in order:
    s = sheets[name]
    if s["type"] == "timeseries":
        print(f"  {name[:34]:34s} yrs {s['years'][0]}-{s['years'][-1]} ({len(s['years'])})  entities={len(s['entities'])}  extra={len(s['extraCols'])}")
